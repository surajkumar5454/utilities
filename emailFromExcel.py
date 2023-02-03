# Enable two-factor authentication for your account....then generate App password and use the generated password here

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import openpyxl
from datetime import datetime
import time

emailCounter = 0
fromaddr = 'ssbsignalcentredli@gmail.com'
#fromaddr = 'controlroomfhqssb@gmail.com'
# fromaddr = 'controlroomssbfhq@gmail.com'
# fromaddr = 'controlroomssbdelhi@gmail.com'
rollNo = ""
name = ""
email = ""

# Loading Excel in Dataframe
dataframe = openpyxl.load_workbook("email_records.xlsx")

# Define variable to read sheet
df = dataframe['SHAMSHI']
file_path = "Recruitment_Letters/Shamshi - 599"

# Iterate the loop to read the cell values
for row in range(0, df.max_row):
    for col in df.iter_cols(1, 1):
        rollNo = col[row].value
    for col in df.iter_cols(2, 2):
        name = col[row].value
    for col in df.iter_cols(3, 3):
        email = col[row].value

    try:

        # instance of MIMEMultipart
        msg = MIMEMultipart()

        # storing the senders email address
        msg['From'] = fromaddr

        # storing the subject
        msg['Subject'] = "Regarding issue of Offer of Appointment from SSB"

        toaddr = email
        msg['To'] = toaddr
        rollNo = str(rollNo)

        body = f'''Dear {name}, 
        
                   Kindly find attached a copy of Offer of Appointment issued to you for the post of Constable(GD) in SSB,
        for further necessary action.
                        
        Best Wishes,
        Sashastra Seema Bal
                '''
        msg.attach(MIMEText(body, 'plain'))

        # open the file to be sent
        filename = "Appointment Letter.pdf"
        attachment = open(file_path + "/" + rollNo + ".pdf", "rb")

        # instance of MIMEBase and named as p
        p = MIMEBase('application', 'octet-stream')

        # To change the payload into encoded form
        p.set_payload(attachment.read())

        # encode into base64
        encoders.encode_base64(p)

        p.add_header('Content-Disposition', "attachment; filename= %s" % filename)

        # attach the instance 'p' to instance 'msg'
        msg.attach(p)

        # creates SMTP session
        s = smtplib.SMTP('smtp.gmail.com', 587)

        # start TLS for security
        s.starttls()

        # Authentication
        s.login(fromaddr, "owyvqzdtgljbvmoy")

        # Converts the Multipart msg into a string
        text = msg.as_string()

        # sending the mail
        # print("before send\n")
        s.sendmail(fromaddr, toaddr, text)
        # print("after send\n")
        emailCounter = emailCounter + 1
        # terminating the session
        s.quit()
        print("Total sent: " + str(emailCounter))
        now = str(datetime.now())
        f1 = open("successLog.txt", "a+")
        f1.write(
            now + f" Email sent successfully to {rollNo}: {name}. File: {file_path}/{rollNo}.pdf   Email : {email}\n")
        f1.close()
        # time.sleep(1)

    except Exception as e:
        print(f"Error occured while sending mail to {name} at S.No {rollNo}")
        print(e)
        now = str(datetime.now())
        f = open("errorLog.txt", "a")
        f.write(now + f"   Error occured while sending mail to {name}  at S.No {rollNo}\n")
        f.write(now + "   " + str(e) + "\n")
        f.close()
