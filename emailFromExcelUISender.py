# Enable two-factor authentication for your account....then generate App password and use the generated password here

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import openpyxl
from datetime import datetime
from tkinter import filedialog
from tkinter.ttk import *
import tkinter as tk
from tkinter import scrolledtext
from PIL import ImageTk, Image

root = tk.Tk()
root.geometry("700x450")
root.title("Email Sender (Sashastra Seema Bal)")


# image2 = Image.open("ssb_logo.png")
# image2 = image2.resize((65, 65), Image.ANTIALIAS)
# test1 = ImageTk.PhotoImage(image2)
# label2 = tk.Label(image=test1)
# label2.image = test1
# label2.place(x=600, y=0)

lbl_blank = Label(text="")
lbl_blank.grid(row=1, column=3)
lbl_blank.grid(row=2, column=3)

lbl = Label(text="Upload Excel File")
lbl.grid(row=3, column=3)
btn = Button(text="Upload", command=lambda: upload_file())
btn.grid(row=3, column=4)
lbl2 = Label(text="Directory containing attachments")
lbl2.grid(row=4, column=3)
btn2 = Button(text="Attachments", command=lambda: setAttachmentPath())
btn2.grid(row=4, column=4)

lbl = Label(text="Enter Subject of Email")
lbl.grid(row=5, column=3)
# Create an Entry widget to accept User Input
subject = Entry(root, width=40, font=("Times New Roman", 15))
subject.focus_set()
subject.grid(row=5, column=4)

lbl = Label(text="Body of Email")
lbl.grid(row=6, column=3)

tbody = scrolledtext.ScrolledText(root, wrap=tk.WORD, width=40, height=8, font=("Times New Roman", 15))
tbody.grid(column=4, row=6, pady=10, padx=10)

# Create an Entry widget to accept User Input
# tbody = Entry(root, width=40)
# tbody.focus_set()
# tbody.grid(row=4, column=4)

lbl = Label(text="Email Address")
lbl.grid(row=7, column=3)
# Create an Entry widget to accept User Input
fromAddress = Entry(root, width=40, font=("Times New Roman", 15))
fromAddress.focus_set()
fromAddress.grid(row=7, column=4)

lbl = Label(text="Password")
lbl.grid(row=8, column=3)
# Create an Entry widget to accept User Input
password = Entry(root, show="*", width=40, font=("Times New Roman", 15))
password.focus_set()
password.grid(row=8, column=4)

lbl_blank = Label(text="")
lbl_blank.grid(row=9, column=3)

btn3 = Button(text="SEND MAIL", command=lambda: sendMail())
btn3.grid(row=11, column=4)

btn2["state"] = "disabled"
btn3["state"] = "disabled"

success_msg = tk.StringVar()
lbl3 = tk.Label(textvariable=success_msg, fg="#256ef5")
lbl3.grid(row=12, column=3)
success_msg.set("")

btn5 = tk.Button(text="CLOSE", command=lambda: close_win())
btn5.grid(row=13, column=4)

itmsg = tk.Label(text="Created By: SOFTWARE CELL, FHQ SSB DELHI ", fg="green")
itmsg.grid(row=13, column=3)


def upload_file():
    local_files = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
    global file
    file = local_files
    if local_files != "":
        btn2["state"] = "enable"
    success_msg.set("")
    return


def close_win():
    root.destroy()


def setAttachmentPath():
    local_savefile = filedialog.askdirectory()
    global attachmentFolderPath
    attachmentFolderPath = local_savefile
    if local_savefile != "":
        btn3["state"] = "enable"
    return


def sendMail():
    # fromaddr = 'surajkumar.geu@gmail.com'
    fromaddr = fromAddress.get()
    passwrd = password.get()
    success_msg.set("Please Wait ...  ")
    root.update()
    rollNo = ""
    name = ""
    email = ""
    emailCounter = 0
    # Loading Excel in Dataframe
    # dataframe = openpyxl.load_workbook("records.xlsx")
    dataframe = openpyxl.load_workbook(file)

    # Define variable to read sheet
    df = dataframe.active
    f1 = open("successLog.txt", "a+")

    # Iterate the loop to read the cell values
    for row in range(0, df.max_row):
        for col in df.iter_cols(1, 1):
            rollNo = col[row].value
        for col in df.iter_cols(2, 2):
            name = col[row].value
        for col in df.iter_cols(3, 3):
            email = col[row].value

        try:

            # instance of MIMEMultipart
            msg = MIMEMultipart()

            # storing the senders email address
            msg['From'] = fromaddr

            # storing the subject
            # msg['Subject'] = "SSB Joining Letter"
            msg['Subject'] = subject.get()
            msgBody = tbody.get('1.0', tk.END)
            toaddr = email
            msg['To'] = toaddr
            rollNo = str(rollNo)

            body = f'''Hi {name}, \n       {msgBody}'''
            msg.attach(MIMEText(body, 'plain'))

            # open the file to be sent
            filename = "Joining Letter.pdf"
            attachment = open(attachmentFolderPath + "/" + rollNo + ".pdf", "rb")

            # instance of MIMEBase and named as p
            p = MIMEBase('application', 'octet-stream')

            # To change the payload into encoded form
            p.set_payload(attachment.read())

            # encode into base64
            encoders.encode_base64(p)

            p.add_header('Content-Disposition', "attachment; filename= %s" % filename)

            # attach the instance 'p' to instance 'msg'
            msg.attach(p)

            # creates SMTP session
            s = smtplib.SMTP('smtp.gmail.com', 587)

            # start TLS for security
            s.starttls()

            # Authentication
            # s.login(fromaddr, "ifvgnghhygahhgdd")
            s.login(fromaddr, passwrd)

            # Converts the Multipart msg into a string
            text = msg.as_string()

            # sending the mail
            s.sendmail(fromaddr, toaddr, text)
            emailCounter = emailCounter + 1
            # terminating the session
            s.quit()
            print("Total sent: " + str(emailCounter))
            success_msg.set("Total sent: " + str(emailCounter))
            root.update()
            now = str(datetime.now())
            f1.write(now + f"   Email sent successfully to {name} at S.No {rollNo}. File: {rollNo}.pdf \n")

        except Exception as e:
            print(f"Error occured while sending mail to {name}")
            print(e)
            now = str(datetime.now())
            f = open("errorLog.txt", "a+")
            f.write(now + f"   Error occured while sending mail to {name} \n")
            f.write(now + "   " + str(e) + "\n")
            f.close()
    success_msg.set("Mail sending complete. Total Emails sent : " + str(emailCounter))
    f1.close()


root.mainloop()
