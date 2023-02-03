# Enable two-factor authentication for your account....then generate App password and use the generated password here

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import openpyxl
from datetime import datetime
from tkinter import filedialog
from tkinter.ttk import *
# import tkinter as tk
from tkinter import scrolledtext
from PIL import ImageTk, Image
import customtkinter as tk

tk.set_appearance_mode("dark")
tk.set_default_color_theme("dark-blue")

root = tk.CTk()
root.geometry("550x450")
root.title("Email Sender (Sashastra Seema Bal)")

frame = tk.CTkFrame(master=root)
frame.pack(pady=40, padx=60, fill="both", expand=True)

lbl = tk.CTkLabel(master=frame, text="Upload Excel File")
lbl.grid(row=3, column=3)
btn = tk.CTkButton(master=frame, text="Upload", command=lambda: upload_file())
btn.grid(row=3, column=4)
lbl2 = tk.CTkLabel(master=frame, text="Directory containing attachments")
lbl2.grid(row=4, column=3)
btn2 = Button(master=frame, text="Attachments", command=lambda: setAttachmentPath())
btn2.grid(row=4, column=4)

lbl = tk.CTkLabel(master=frame, text="Enter Subject of Email", width=100)
lbl.grid(row=5, column=3)
# Create an Entry widget to accept User Input
subject = tk.CTkEntry(master=frame, width=40)
subject.focus_set()
subject.grid(row=5, column=4)

lbl = tk.CTkLabel(master=frame, text="Body of Email")
lbl.grid(row=6, column=3)

tbody = tk.CTkEntry(master=frame, width=40)
tbody.focus_set()
tbody.grid(row=6, column=4)

lbl = tk.CTkLabel(master=frame, text="Email Address")
lbl.grid(row=7, column=3)
# Create an Entry widget to accept User Input
fromAddress = tk.CTkEntry(master=frame, width=40)
fromAddress.focus_set()
fromAddress.grid(row=7, column=4)

lbl = tk.CTkLabel(master=frame, text="Password")
lbl.grid(row=8, column=3)
# Create an Entry widget to accept User Input
password = tk.CTkEntry(master=frame, show="*", width=40)
password.focus_set()
password.grid(row=8, column=4)

lbl_blank = tk.CTkLabel(master=frame, text="")
lbl_blank.grid(row=9, column=3)

btn3 = tk.CTkButton(master=frame, text="SEND MAIL", command=lambda: sendMail())
btn3.grid(row=10, column=4)

btn2["state"] = "disabled"
btn3["state"] = "disabled"

btn5 = tk.CTkButton(master=frame, text="CLOSE", command=lambda: close_win())
btn5.grid(row=11, column=4)


def upload_file():
    local_files = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
    global file
    file = local_files
    if local_files != "":
        btn2["state"] = "enable"
    return


def close_win():
    root.destroy()


def setAttachmentPath():
    local_savefile = filedialog.askdirectory()
    global attachmentFolderPath
    attachmentFolderPath = local_savefile
    if local_savefile != "":
        btn3["state"] = "enable"
    return


def sendMail():
    # fromaddr = 'surajkumar.geu@gmail.com'
    fromaddr = fromAddress.get()
    passwrd = password.get()
    # success_msg.set("Please Wait ...  ")
    root.update()
    rollNo = ""
    name = ""
    email = ""
    emailCounter = 0
    # Loading Excel in Dataframe
    # dataframe = openpyxl.load_workbook("records.xlsx")
    dataframe = openpyxl.load_workbook(file)

    # Define variable to read sheet
    df = dataframe.active
    f1 = open("successLog.txt", "a+")

    # Iterate the loop to read the cell values
    for row in range(0, df.max_row):
        for col in df.iter_cols(1, 1):
            rollNo = col[row].value
        for col in df.iter_cols(2, 2):
            name = col[row].value
        for col in df.iter_cols(3, 3):
            email = col[row].value

        try:

            # instance of MIMEMultipart
            msg = MIMEMultipart()

            # storing the senders email address
            msg['From'] = fromaddr

            # storing the subject
            # msg['Subject'] = "SSB Joining Letter"
            msg['Subject'] = subject.get()
            msgBody = tbody.get('1.0', tk.END)
            toaddr = email
            msg['To'] = toaddr
            rollNo = str(rollNo)

            body = f'''Hi {name}, \n       {msgBody}'''
            msg.attach(MIMEText(body, 'plain'))

            # open the file to be sent
            filename = "Joining Letter.pdf"
            attachment = open(attachmentFolderPath + "/" + rollNo + ".pdf", "rb")

            # instance of MIMEBase and named as p
            p = MIMEBase('application', 'octet-stream')

            # To change the payload into encoded form
            p.set_payload(attachment.read())

            # encode into base64
            encoders.encode_base64(p)

            p.add_header('Content-Disposition', "attachment; filename= %s" % filename)

            # attach the instance 'p' to instance 'msg'
            msg.attach(p)

            # creates SMTP session
            s = smtplib.SMTP('smtp.gmail.com', 587)

            # start TLS for security
            s.starttls()

            # Authentication
            # s.login(fromaddr, "ifvgnghhygahhgdd")
            s.login(fromaddr, passwrd)

            # Converts the Multipart msg into a string
            text = msg.as_string()

            # sending the mail
            s.sendmail(fromaddr, toaddr, text)
            emailCounter = emailCounter + 1
            # terminating the session
            s.quit()
            print("Total sent: " + str(emailCounter))
            # success_msg.set("Total sent: " + str(emailCounter))
            root.update()
            now = str(datetime.now())
            f1.write(now + f"   Email sent successfully to {name} at S.No {rollNo}. File: {rollNo}.pdf \n")

        except Exception as e:
            print(f"Error occured while sending mail to {name}")
            print(e)
            now = str(datetime.now())
            f = open("errorLog.txt", "a+")
            f.write(now + f"   Error occured while sending mail to {name} \n")
            f.write(now + "   " + str(e) + "\n")
            f.close()
    # success_msg.set("Mail sending complete. Total Emails sent : " + str(emailCounter))
    f1.close()


root.mainloop()
