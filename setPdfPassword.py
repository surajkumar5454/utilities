import PyPDF2
import openpyxl

# Loading Excel in Dataframe
dataframe = openpyxl.load_workbook("records.xlsx")

# Define variable to read sheet
df = dataframe.active

# Iterate the loop to read the cell values
for row in range(0, df.max_row):
    for col in df.iter_cols(1, 1):
        rollNo = col[row].value
    for col in df.iter_cols(2, 2):
        name = col[row].value
    for col in df.iter_cols(3, 3):
        email = col[row].value
    for col in df.iter_cols(4, 4):
        passwd = str(col[row].value)
    pdf_in_file = open(f"withoutpassword/{rollNo}.pdf", 'rb')
    inputpdf = PyPDF2.PdfFileReader(pdf_in_file)
    pages_no = inputpdf.numPages
    output = PyPDF2.PdfFileWriter()

    for i in range(pages_no):
        inputpdf = PyPDF2.PdfFileReader(pdf_in_file)

        output.addPage(inputpdf.getPage(i))
        output.encrypt(passwd)

        # with open("simple_password_protected.pdf", "wb") as outputStream:
        with open(f"passwordpdf/{rollNo}.pdf", "wb") as outputStream:
            output.write(outputStream)

    pdf_in_file.close()
